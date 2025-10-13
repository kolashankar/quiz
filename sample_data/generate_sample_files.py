import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# GATE EEE Questions - CSV Format
gate_eee_questions = [
    {
        "sub_section_id": "sample_subsection_1",
        "question_text": "A 3-phase, 50 Hz, 4-pole induction motor runs at 1440 RPM at full load. What is the slip?",
        "option1": "2%",
        "option2": "4%",
        "option3": "6%",
        "option4": "8%",
        "correct_answer": 1,
        "difficulty": "easy",
        "tags": "Electrical Machines, Induction Motor",
        "explanation": "Synchronous speed Ns = (120 × f) / P = (120 × 50) / 4 = 1500 RPM. Slip s = (Ns - N) / Ns = (1500 - 1440) / 1500 = 0.04 = 4%",
        "hint": "First calculate the synchronous speed using the formula Ns = (120 × f) / P",
        "solution": "Step 1: Calculate synchronous speed\nNs = (120 × f) / P = (120 × 50) / 4 = 1500 RPM\n\nStep 2: Calculate slip\ns = (Ns - N) / Ns = (1500 - 1440) / 1500 = 60/1500 = 0.04 = 4%",
        "code_snippet": "",
        "image_url": "",
        "formula": "N_s = \\frac{120 \\times f}{P}, s = \\frac{N_s - N}{N_s}"
    },
    {
        "sub_section_id": "sample_subsection_1",
        "question_text": "In a transformer, the primary winding has 1000 turns and secondary has 200 turns. If the primary voltage is 220V, what is the secondary voltage?",
        "option1": "22V",
        "option2": "44V",
        "option3": "66V",
        "option4": "88V",
        "correct_answer": 1,
        "difficulty": "easy",
        "tags": "Transformer, Electrical Machines",
        "explanation": "Transformation ratio K = N2/N1 = 200/1000 = 0.2. Secondary voltage V2 = K × V1 = 0.2 × 220 = 44V",
        "hint": "Use the transformer turns ratio formula V2/V1 = N2/N1",
        "solution": "Given:\nPrimary turns N1 = 1000\nSecondary turns N2 = 200\nPrimary voltage V1 = 220V\n\nTransformation ratio: K = N2/N1 = 200/1000 = 0.2\nSecondary voltage: V2 = K × V1 = 0.2 × 220 = 44V",
        "code_snippet": "",
        "image_url": "",
        "formula": "\\frac{V_2}{V_1} = \\frac{N_2}{N_1}"
    },
    {
        "sub_section_id": "sample_subsection_1",
        "question_text": "A DC shunt motor runs at 1000 RPM when the armature current is 50A. If the armature resistance is 0.5Ω and the supply voltage is 250V, find the back EMF.",
        "option1": "200V",
        "option2": "225V",
        "option3": "235V",
        "option4": "240V",
        "correct_answer": 1,
        "difficulty": "medium",
        "tags": "DC Motor, Electrical Machines",
        "explanation": "Back EMF Eb = V - Ia × Ra = 250 - (50 × 0.5) = 250 - 25 = 225V",
        "hint": "Back EMF can be calculated by subtracting the voltage drop across armature resistance from supply voltage",
        "solution": "Given:\nSupply voltage V = 250V\nArmature current Ia = 50A\nArmature resistance Ra = 0.5Ω\n\nBack EMF:\nEb = V - Ia × Ra\nEb = 250 - (50 × 0.5)\nEb = 250 - 25\nEb = 225V",
        "code_snippet": "",
        "image_url": "",
        "formula": "E_b = V - I_a \\times R_a"
    },
    {
        "sub_section_id": "sample_subsection_1",
        "question_text": "A transmission line has a series impedance of (4 + j6)Ω and shunt admittance of j0.001S. Calculate the magnitude of characteristic impedance.",
        "option1": "77.46Ω",
        "option2": "87.46Ω",
        "option3": "97.46Ω",
        "option4": "107.46Ω",
        "correct_answer": 0,
        "difficulty": "hard",
        "tags": "Transmission Lines, Power Systems",
        "explanation": "Characteristic impedance Zc = √(Z/Y) = √((4+j6)/(j0.001)) = √(7.211/0.001∠90°) = √7211∠-90° = 84.92∠-45° ≈ 77.46Ω",
        "hint": "Use the formula Zc = √(Z/Y) where Z is series impedance and Y is shunt admittance",
        "solution": "Given:\nSeries impedance Z = 4 + j6 Ω\nMagnitude |Z| = √(4² + 6²) = √52 = 7.211Ω\nShunt admittance Y = j0.001 S\n\nCharacteristic impedance:\nZc = √(Z/Y)\nZc = √(7.211∠56.31° / 0.001∠90°)\nZc = √(7211∠-33.69°)\nZc = 84.92∠-16.85°\n|Zc| ≈ 77.46Ω",
        "code_snippet": "",
        "image_url": "",
        "formula": "Z_c = \\sqrt{\\frac{Z}{Y}}"
    },
    {
        "sub_section_id": "sample_subsection_1",
        "question_text": "In a power system, the per unit impedance of a transformer is 0.1pu. If the base power is 100MVA and base voltage is 11kV, what is the actual impedance in ohms?",
        "option1": "0.121Ω",
        "option2": "1.21Ω",
        "option3": "12.1Ω",
        "option4": "121Ω",
        "correct_answer": 0,
        "difficulty": "medium",
        "tags": "Per Unit System, Power Systems",
        "explanation": "Base impedance Zbase = V²base/Sbase = (11×10³)²/(100×10⁶) = 1.21Ω. Actual impedance = Zpu × Zbase = 0.1 × 1.21 = 0.121Ω",
        "hint": "First calculate base impedance using Zbase = V²base/Sbase",
        "solution": "Given:\nPer unit impedance Zpu = 0.1 pu\nBase power Sbase = 100 MVA = 100 × 10⁶ VA\nBase voltage Vbase = 11 kV = 11 × 10³ V\n\nStep 1: Calculate base impedance\nZbase = V²base / Sbase\nZbase = (11 × 10³)² / (100 × 10⁶)\nZbase = 121 × 10⁶ / 100 × 10⁶\nZbase = 1.21Ω\n\nStep 2: Calculate actual impedance\nZactual = Zpu × Zbase = 0.1 × 1.21 = 0.121Ω",
        "code_snippet": "",
        "image_url": "",
        "formula": "Z_{base} = \\frac{V_{base}^2}{S_{base}}, Z_{actual} = Z_{pu} \\times Z_{base}"
    },
    {
        "sub_section_id": "sample_subsection_1",
        "question_text": "A three-phase star-connected alternator has a line voltage of 11kV. What is the phase voltage?",
        "option1": "5.5kV",
        "option2": "6.35kV",
        "option3": "7.8kV",
        "option4": "11kV",
        "correct_answer": 1,
        "difficulty": "easy",
        "tags": "AC Systems, Electrical Machines",
        "explanation": "In star connection, Phase voltage Vph = VL/√3 = 11/√3 = 11/1.732 ≈ 6.35kV",
        "hint": "In star connection, line voltage is √3 times the phase voltage",
        "solution": "Given:\nLine voltage VL = 11 kV\n\nFor star connection:\nVph = VL / √3\nVph = 11 / 1.732\nVph ≈ 6.35 kV",
        "code_snippet": "",
        "image_url": "",
        "formula": "V_{ph} = \\frac{V_L}{\\sqrt{3}}"
    },
    {
        "sub_section_id": "sample_subsection_1",
        "question_text": "A circuit breaker has a recovery voltage of 132kV. If the rate of rise of recovery voltage (RRRV) is 2kV/μs, what is the total time for voltage recovery?",
        "option1": "44μs",
        "option2": "56μs",
        "option3": "66μs",
        "option4": "76μs",
        "correct_answer": 2,
        "difficulty": "medium",
        "tags": "Circuit Breakers, Switchgear",
        "explanation": "Time for voltage recovery = Recovery voltage / RRRV = 132 / 2 = 66μs",
        "hint": "Divide the recovery voltage by the rate of rise of recovery voltage",
        "solution": "Given:\nRecovery voltage = 132 kV\nRRRV = 2 kV/μs\n\nTime for voltage recovery:\nTime = Recovery voltage / RRRV\nTime = 132 / 2\nTime = 66 μs",
        "code_snippet": "",
        "image_url": "",
        "formula": "t = \\frac{V_{recovery}}{RRRV}"
    },
    {
        "sub_section_id": "sample_subsection_1",
        "question_text": "A single-phase transformer has a rating of 10kVA, 2200/220V. What is the rated primary current?",
        "option1": "4.55A",
        "option2": "5.45A",
        "option3": "6.35A",
        "option4": "7.25A",
        "correct_answer": 0,
        "difficulty": "easy",
        "tags": "Transformer, Electrical Machines",
        "explanation": "Primary current I1 = kVA × 1000 / V1 = 10 × 1000 / 2200 = 4.55A",
        "hint": "Use the power formula S = V × I to find current",
        "solution": "Given:\nTransformer rating S = 10 kVA = 10,000 VA\nPrimary voltage V1 = 2200 V\n\nPrimary current:\nI1 = S / V1\nI1 = 10,000 / 2200\nI1 ≈ 4.55 A",
        "code_snippet": "",
        "image_url": "",
        "formula": "I_1 = \\frac{S}{V_1}"
    },
    {
        "sub_section_id": "sample_subsection_1",
        "question_text": "In a balanced 3-phase system, the red phase voltage is 230∠0°V. What is the yellow phase voltage in a RYB sequence?",
        "option1": "230∠120°V",
        "option2": "230∠-120°V",
        "option3": "230∠240°V",
        "option4": "230∠-240°V",
        "correct_answer": 1,
        "difficulty": "medium",
        "tags": "3-Phase Systems, AC Circuits",
        "explanation": "In RYB sequence, yellow phase lags red phase by 120°. So VY = 230∠(0°-120°) = 230∠-120°V",
        "hint": "In a balanced 3-phase system with RYB sequence, each phase is separated by 120°",
        "solution": "Given:\nRed phase voltage VR = 230∠0° V\nPhase sequence: RYB\n\nIn RYB sequence:\n- Red (R) phase: 0°\n- Yellow (Y) phase: -120° (lags by 120°)\n- Blue (B) phase: -240° or +120°\n\nYellow phase voltage:\nVY = 230∠-120° V",
        "code_snippet": "",
        "image_url": "",
        "formula": "V_Y = V_R \\angle -120^\\circ"
    },
    {
        "sub_section_id": "sample_subsection_1",
        "question_text": "A protective relay has a time multiplier setting (TMS) of 0.3 and operates with a plug setting of 150%. If the fault current is 3000A and CT ratio is 1000/5, what is the operating time? (Use IDMT characteristic)",
        "option1": "0.6s",
        "option2": "0.9s",
        "option3": "1.2s",
        "option4": "1.5s",
        "correct_answer": 1,
        "difficulty": "hard",
        "tags": "Protection, Relays",
        "explanation": "Plug setting multiplier PSM = Fault current / (CT ratio × Plug setting) = 3000 / (1000/5 × 1.5) = 3000 / 300 = 10. For IDMT relay, t = TMS × [0.14 / (PSM^0.02 - 1)] ≈ 0.3 × 3 = 0.9s",
        "hint": "First calculate the plug setting multiplier (PSM), then use IDMT characteristic equation",
        "solution": "Given:\nTMS = 0.3\nPlug setting = 150% = 1.5\nFault current If = 3000 A\nCT ratio = 1000/5 = 200\n\nStep 1: Calculate plug setting current\nIplug = (1000/5) × 1.5 = 200 × 1.5 = 300 A\n\nStep 2: Calculate PSM\nPSM = If / Iplug = 3000 / 300 = 10\n\nStep 3: Calculate operating time (using simplified IDMT)\nt = TMS × [0.14 / (PSM^0.02 - 1)]\nt ≈ 0.3 × 3\nt ≈ 0.9 s",
        "code_snippet": "",
        "image_url": "",
        "formula": "PSM = \\frac{I_f}{I_{plug}}, t = TMS \\times \\frac{0.14}{PSM^{0.02} - 1}"
    }
]

# JEE Advanced Questions - Excel Format
jee_advanced_questions = [
    {
        "sub_section_id": "sample_subsection_2",
        "question_text": "If f(x) = x³ - 3x + 2, find the number of real roots of the equation f(f(x)) = 0",
        "option1": "5",
        "option2": "7",
        "option3": "9",
        "option4": "11",
        "correct_answer": 1,
        "difficulty": "hard",
        "tags": "Calculus, Functions",
        "explanation": "First solve f(x) = 0 to get roots, then solve f(x) = each root. The function has 3 real roots for f(x) = 0, and solving f(x) = root for each gives 7 total roots.",
        "hint": "First find roots of f(x) = 0, then solve f(x) = r for each root r",
        "solution": "Step 1: Find roots of f(x) = 0\nx³ - 3x + 2 = 0\n(x-1)²(x+2) = 0\nRoots: x = 1 (double), x = -2\n\nStep 2: Solve f(f(x)) = 0\nThis means f(x) must equal one of the roots: 1, 1, or -2\n\nSolve f(x) = 1: x³ - 3x + 2 = 1 → x³ - 3x + 1 = 0 (3 real roots)\nSolve f(x) = -2: x³ - 3x + 2 = -2 → x³ - 3x + 4 = 0 (1 real root)\n\nTotal: 3 + 1 + 3 = 7 real roots",
        "code_snippet": "",
        "image_url": "",
        "formula": "f(x) = x^3 - 3x + 2"
    },
    {
        "sub_section_id": "sample_subsection_2",
        "question_text": "In a triangle ABC, if sin²A + sin²B + sin²C = 2, then the triangle is:",
        "option1": "Right-angled",
        "option2": "Equilateral",
        "option3": "Isosceles",
        "option4": "Scalene",
        "correct_answer": 0,
        "difficulty": "medium",
        "tags": "Trigonometry, Geometry",
        "explanation": "Given sin²A + sin²B + sin²C = 2. For a triangle, A + B + C = π. If triangle is right-angled, say C = π/2, then sin²A + sin²B + 1 = 2, which gives sin²A + sin²B = 1, satisfied when A + B = π/2.",
        "hint": "Check if the condition is satisfied for a right-angled triangle",
        "solution": "Given: sin²A + sin²B + sin²C = 2\n\nFor a right-angled triangle with C = 90°:\nsinC = 1, so sin²C = 1\n\nThen: sin²A + sin²B + 1 = 2\nsin²A + sin²B = 1\n\nSince A + B = 90° (as C = 90° and A + B + C = 180°)\nsin²A + sin²B = sin²A + cos²A = 1 ✓\n\nTherefore, the triangle is right-angled.",
        "code_snippet": "",
        "image_url": "",
        "formula": "\\sin^2 A + \\sin^2 B + \\sin^2 C = 2"
    },
    {
        "sub_section_id": "sample_subsection_2",
        "question_text": "The number of ways in which 5 boys and 3 girls can be seated in a row so that no two girls sit together is:",
        "option1": "14400",
        "option2": "28800",
        "option3": "43200",
        "option4": "57600",
        "correct_answer": 0,
        "difficulty": "medium",
        "tags": "Permutations, Combinatorics",
        "explanation": "First arrange 5 boys in 5! ways = 120. This creates 6 gaps (X_B_X_B_X_B_X_B_X_B_X). Select 3 gaps from 6 for girls: ⁶C₃ = 20. Arrange 3 girls: 3! = 6. Total = 120 × 20 × 6 = 14400",
        "hint": "First arrange boys, then place girls in the gaps between boys",
        "solution": "Step 1: Arrange 5 boys in a row\nNumber of ways = 5! = 120\n\nStep 2: After arranging boys, there are 6 positions for girls (before first boy, between boys, and after last boy):\n_B_B_B_B_B_\n\nStep 3: Choose 3 positions from 6 for girls\nNumber of ways = ⁶C₃ = 6!/(3!×3!) = 20\n\nStep 4: Arrange 3 girls in chosen positions\nNumber of ways = 3! = 6\n\nTotal arrangements = 120 × 20 × 6 = 14,400",
        "code_snippet": "",
        "image_url": "",
        "formula": "^nC_r = \\frac{n!}{r!(n-r)!}"
    },
    {
        "sub_section_id": "sample_subsection_2",
        "question_text": "If the coefficient of x⁷ in (ax² + 1/bx)¹¹ equals the coefficient of x⁻⁷ in (ax - 1/bx²)¹¹, then ab equals:",
        "option1": "1",
        "option2": "-1",
        "option3": "0",
        "option4": "2",
        "correct_answer": 0,
        "difficulty": "hard",
        "tags": "Binomial Theorem, Algebra",
        "explanation": "For (ax² + 1/bx)¹¹, x⁷ term: ¹¹C₄(ax²)⁷(1/bx)⁴. For (ax - 1/bx²)¹¹, x⁻⁷ term: ¹¹C₇(ax)⁴(-1/bx²)⁷. Equating coefficients gives ab = 1",
        "hint": "Use binomial theorem to find the coefficients and equate them",
        "solution": "For (ax² + 1/bx)¹¹:\nGeneral term: Tr+1 = ¹¹Cr(ax²)^(11-r)(1/bx)^r\nPower of x: 2(11-r) - r = 22 - 3r\nFor x⁷: 22 - 3r = 7 → r = 5\nCoefficient: ¹¹C₅ × a⁶ × b⁻⁵\n\nFor (ax - 1/bx²)¹¹:\nGeneral term: Tr+1 = ¹¹Cr(ax)^(11-r)(-1/bx²)^r\nPower of x: (11-r) - 2r = 11 - 3r\nFor x⁻⁷: 11 - 3r = -7 → r = 6\nCoefficient: ¹¹C₆ × a⁵ × (-1)⁶ × b⁻⁶\n\nEquating:\n¹¹C₅ × a⁶ × b⁻⁵ = ¹¹C₆ × a⁵ × b⁻⁶\nSince ¹¹C₅ = ¹¹C₆:\na⁶ × b⁻⁵ = a⁵ × b⁻⁶\na × b = 1\nab = 1",
        "code_snippet": "",
        "image_url": "",
        "formula": "(a+b)^n = \\sum_{r=0}^{n} ^nC_r \\cdot a^{n-r} \\cdot b^r"
    },
    {
        "sub_section_id": "sample_subsection_2",
        "question_text": "The general solution of the differential equation dy/dx + y = e⁻ˣ is:",
        "option1": "y = (x + c)e⁻ˣ",
        "option2": "y = (x - c)e⁻ˣ",
        "option3": "y = (cx + 1)e⁻ˣ",
        "option4": "y = (c - x)e⁻ˣ",
        "correct_answer": 0,
        "difficulty": "medium",
        "tags": "Differential Equations, Calculus",
        "explanation": "This is a first-order linear DE. IF = e∫¹dx = eˣ. Solution: y×eˣ = ∫e⁻ˣ×eˣdx = ∫1dx = x + c. Therefore y = (x + c)e⁻ˣ",
        "hint": "This is a first-order linear differential equation. Find the integrating factor",
        "solution": "Given: dy/dx + y = e⁻ˣ\n\nThis is of the form: dy/dx + Py = Q\nwhere P = 1, Q = e⁻ˣ\n\nStep 1: Find integrating factor (IF)\nIF = e^∫P dx = e^∫1 dx = eˣ\n\nStep 2: Multiply equation by IF\neˣ(dy/dx) + eˣy = eˣ×e⁻ˣ\nd/dx(yeˣ) = 1\n\nStep 3: Integrate both sides\nyeˣ = ∫1 dx = x + c\n\nStep 4: Solve for y\ny = (x + c)e⁻ˣ",
        "code_snippet": "",
        "image_url": "",
        "formula": "IF = e^{\\int P dx}"
    },
    {
        "sub_section_id": "sample_subsection_2",
        "question_text": "In an experiment, two identical balls are dropped from a height h. One ball is given horizontal velocity v₀. What is the ratio of their times of flight?",
        "option1": "1:1",
        "option2": "1:2",
        "option3": "2:1",
        "option4": "√2:1",
        "correct_answer": 0,
        "difficulty": "easy",
        "tags": "Projectile Motion, Physics",
        "explanation": "Time of flight depends only on vertical motion: t = √(2h/g). Horizontal velocity doesn't affect vertical motion. So both balls have same time of flight. Ratio = 1:1",
        "hint": "Time of flight in projectile motion depends only on vertical component of motion",
        "solution": "For vertical motion (both balls):\nInitial vertical velocity u = 0\nHeight h\nAcceleration = g (downward)\n\nUsing equation: h = ut + (1/2)gt²\nh = 0 + (1/2)gt²\nt = √(2h/g)\n\nHorizontal velocity doesn't affect vertical motion.\nBoth balls have same vertical motion parameters.\n\nTherefore, time of flight ratio = 1:1",
        "code_snippet": "",
        "image_url": "",
        "formula": "t = \\sqrt{\\frac{2h}{g}}"
    },
    {
        "sub_section_id": "sample_subsection_2",
        "question_text": "A parallel plate capacitor with plate area A and separation d is filled with a dielectric of constant K. What is its capacitance?",
        "option1": "Kε₀A/d",
        "option2": "ε₀A/Kd",
        "option3": "KAd/ε₀",
        "option4": "ε₀/KAd",
        "correct_answer": 0,
        "difficulty": "easy",
        "tags": "Electrostatics, Capacitance",
        "explanation": "Capacitance of parallel plate capacitor with dielectric: C = Kε₀A/d, where K is dielectric constant",
        "hint": "Dielectric constant K multiplies the capacitance of empty capacitor",
        "solution": "For parallel plate capacitor:\n\nWithout dielectric:\nC₀ = ε₀A/d\n\nWith dielectric of constant K:\nC = KC₀\nC = K × (ε₀A/d)\nC = Kε₀A/d\n\nwhere:\nK = dielectric constant\nε₀ = permittivity of free space\nA = plate area\nd = separation between plates",
        "code_snippet": "",
        "image_url": "",
        "formula": "C = \\frac{K\\varepsilon_0 A}{d}"
    },
    {
        "sub_section_id": "sample_subsection_2",
        "question_text": "The de Broglie wavelength associated with an electron accelerated through a potential difference V is:",
        "option1": "λ = h/√(2meV)",
        "option2": "λ = h√(2meV)",
        "option3": "λ = √(h/2meV)",
        "option4": "λ = 2h/√(meV)",
        "correct_answer": 0,
        "difficulty": "medium",
        "tags": "Modern Physics, Quantum Mechanics",
        "explanation": "Kinetic energy KE = eV. Momentum p = √(2mKE) = √(2meV). de Broglie wavelength λ = h/p = h/√(2meV)",
        "hint": "Use relation between kinetic energy and potential, then apply de Broglie equation",
        "solution": "When electron is accelerated through potential V:\n\nStep 1: Find kinetic energy\nKE = eV (e = charge of electron)\n\nStep 2: Find momentum\nKE = p²/2m\np² = 2m×KE = 2m×eV\np = √(2meV)\n\nStep 3: Apply de Broglie equation\nλ = h/p\nλ = h/√(2meV)\n\nwhere:\nh = Planck's constant\nm = mass of electron\ne = charge of electron\nV = potential difference",
        "code_snippet": "",
        "image_url": "",
        "formula": "\\lambda = \\frac{h}{p} = \\frac{h}{\\sqrt{2meV}}"
    },
    {
        "sub_section_id": "sample_subsection_2",
        "question_text": "For the reaction 2A + B → 3C, if the rate of disappearance of A is 0.6 mol/L/s, what is the rate of appearance of C?",
        "option1": "0.3 mol/L/s",
        "option2": "0.6 mol/L/s",
        "option3": "0.9 mol/L/s",
        "option4": "1.2 mol/L/s",
        "correct_answer": 2,
        "difficulty": "easy",
        "tags": "Chemical Kinetics, Chemistry",
        "explanation": "Rate = -(1/2)(d[A]/dt) = -(d[B]/dt) = (1/3)(d[C]/dt). Given d[A]/dt = -0.6, so rate = 0.3. Therefore d[C]/dt = 3 × 0.3 = 0.9 mol/L/s",
        "hint": "Rate of reaction is related to stoichiometric coefficients",
        "solution": "Given reaction: 2A + B → 3C\nRate of disappearance of A = 0.6 mol/L/s\n\nFor the reaction aA + bB → cC:\nRate = -(1/a)(d[A]/dt) = -(1/b)(d[B]/dt) = (1/c)(d[C]/dt)\n\nFor our reaction:\nRate = -(1/2)(d[A]/dt) = (1/3)(d[C]/dt)\n\nGiven: -d[A]/dt = 0.6 mol/L/s\n\nStep 1: Find rate\nRate = (1/2) × 0.6 = 0.3 mol/L/s\n\nStep 2: Find d[C]/dt\nd[C]/dt = 3 × Rate = 3 × 0.3 = 0.9 mol/L/s",
        "code_snippet": "",
        "image_url": "",
        "formula": "Rate = -\\frac{1}{a}\\frac{d[A]}{dt} = \\frac{1}{c}\\frac{d[C]}{dt}"
    },
    {
        "sub_section_id": "sample_subsection_2",
        "question_text": "The pH of 0.001M HCl solution is approximately:",
        "option1": "2",
        "option2": "3",
        "option3": "4",
        "option4": "5",
        "correct_answer": 1,
        "difficulty": "easy",
        "tags": "Acids and Bases, Chemistry",
        "explanation": "HCl is a strong acid, completely dissociates. [H⁺] = 0.001M = 10⁻³M. pH = -log[H⁺] = -log(10⁻³) = 3",
        "hint": "HCl is a strong acid that completely dissociates. Use pH = -log[H⁺]",
        "solution": "Given: HCl concentration = 0.001 M = 10⁻³ M\n\nHCl is a strong acid, so it completely dissociates:\nHCl → H⁺ + Cl⁻\n\nTherefore:\n[H⁺] = [HCl] = 0.001 M = 10⁻³ M\n\nCalculate pH:\npH = -log₁₀[H⁺]\npH = -log₁₀(10⁻³)\npH = 3",
        "code_snippet": "",
        "image_url": "",
        "formula": "pH = -\\log_{10}[H^+]"
    }
]

# Create CSV file for GATE EEE
gate_df = pd.DataFrame(gate_eee_questions)
gate_df.to_csv('/app/sample_data/GATE_EEE_Questions.csv', index=False)
print("✓ Created GATE_EEE_Questions.csv")

# Create Excel file for JEE Advanced with formatting
jee_df = pd.DataFrame(jee_advanced_questions)

# Create Excel writer
with pd.ExcelWriter('/app/sample_data/JEE_Advanced_Questions.xlsx', engine='openpyxl') as writer:
    jee_df.to_excel(writer, sheet_name='Questions', index=False)
    
    # Get the workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Questions']
    
    # Style the header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Adjust column widths
    column_widths = {
        'A': 20,  # sub_section_id
        'B': 60,  # question_text
        'C': 30,  # option1
        'D': 30,  # option2
        'E': 30,  # option3
        'F': 30,  # option4
        'G': 15,  # correct_answer
        'H': 12,  # difficulty
        'I': 30,  # tags
        'J': 60,  # explanation
        'K': 50,  # hint
        'L': 70,  # solution
        'M': 50,  # code_snippet
        'N': 30,  # image_url
        'O': 50,  # formula
    }
    
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width
    
    # Set row height and wrap text for all data rows
    for row in worksheet.iter_rows(min_row=2, max_row=len(jee_advanced_questions) + 1):
        worksheet.row_dimensions[row[0].row].height = 60
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

print("✓ Created JEE_Advanced_Questions.xlsx with formatting")
print("\nSample files created successfully!")
print("Location: /app/sample_data/")
print("Files:")
print("  - GATE_EEE_Questions.csv (10 questions)")
print("  - JEE_Advanced_Questions.xlsx (10 questions)")
